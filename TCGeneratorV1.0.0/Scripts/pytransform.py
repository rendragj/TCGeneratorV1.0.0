import sys
import traceback
from unittest.util import _MAX_LENGTH
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time

def read_file_with_fallback(file_path):
    encodings = ['utf-8', 'latin1', 'iso-8859-1']
    extension = os.path.splitext(file_path)[1].lower()

    if extension == '.csv':
        for enc in encodings:
            try:
                return pd.read_csv(file_path, encoding=enc, on_bad_lines='skip')
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue
        raise Exception(f"Failed to read {file_path} with all attempted encodings.")
    
    elif extension in ['.xls', '.xlsx']:
        try:
            return pd.read_excel(file_path)
        except Exception as e:
            raise Exception(f"Error reading Excel file {file_path}: {e}")

    else:
        raise Exception(f"Unsupported file type: {file_path}")

def report_progress(progress_file, message, percentage):
    retries = 5
    while retries > 0:
        try:
            with open(progress_file, 'w') as f:
                f.write(f"{percentage}\n{message}")
            break
        except PermissionError:
            retries -= 1
            time.sleep(0.1)
            if retries == 0:
                raise

def preprocess_add_file(add_df):
    add_df.columns = add_df.columns.str.strip()
    add_df['Name'] = add_df['Name'].str.strip().str.lower()
    add_df['Number'] = add_df['Number'].astype(str).str.strip()
    add_df['Condition'] = add_df['Condition'].str.strip().str.lower()
    add_df['Foil Type'] = add_df['Foil Type'].str.strip().str.lower()
    return add_df[['Name', 'Number', 'Qty', 'Foil Type', 'Condition']]

def construct_concat(name, number, condition, foil_type):
    condition_mapping = {
        'mnm': 'Near Mint',
        'lp': 'Lightly Played',
        'mp': 'Moderately Played',
        'hp': 'Heavily Played',
        'dm': 'Damaged'
    }
    condition_str = condition_mapping.get(condition, '')
    foil_str = 'Holofoil' if foil_type != 'normal' else ''
    return f"{name} -{number}-{condition_str} {foil_str}".strip()

def transform_data(add_df, master_df):
    add_df = preprocess_add_file(add_df)
    output_df = pd.DataFrame(columns=master_df.columns)
    not_found_indices = []

    for index, row in add_df.iterrows():
        name = row['Name']
        number = row['Number']
        condition = row['Condition']
        foil_type = row['Foil Type']
        qty = row['Qty']  # Get the quantity from the add file

        concat_str = construct_concat(name, number, condition, foil_type)
        matched_row = master_df[master_df['Concat'].str.lower() == concat_str.lower()]

        if not matched_row.empty:
            matched_row = matched_row.copy()  # Create a copy to modify
            matched_row['Add to Quantity'] = qty  # Update the 'Add to Quantity' column with qty
            output_df = pd.concat([output_df, matched_row], ignore_index=True)
        else:
            not_found_indices.append(len(output_df))
            row_to_add = pd.Series({col: '' for col in output_df.columns})
            for col in ['Name', 'Number', 'Qty', 'Foil Type', 'Condition']:
                row_to_add[col] = row[col] if col in row else ''
            row_to_add['Add to Quantity'] = qty  # Set the 'Add to Quantity' column with qty
            output_df = pd.concat([output_df, row_to_add.to_frame().T], ignore_index=True)

    return output_df, not_found_indices

def highlight_missing_rows(output_file_path, not_found_indices):
    try:
        wb = load_workbook(output_file_path)
        ws = wb.active
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for row_idx in not_found_indices:
            for cell in ws[row_idx + 2]:  # Adding 2 to account for zero-indexing and header row
                cell.fill = red_fill

        wb.save(output_file_path)
        wb.close()
    except Exception as e:
        print(f"Error while highlighting missing rows: {e}")
        traceback.print_exc()
        
def adjust_column_widths(output_file_path):
    try:
        wb = load_workbook(output_file_path)
        ws = wb.active

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    # Calculate the length of the cell's value
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = max_length + 2  # Add some padding
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file_path)
    except Exception as e:
        print(f"Error while adjusting column widths: {e}")
        traceback.print_exc()

def main():
    try:
        if len(sys.argv) != 5:
            print("Usage: python pytransform.py <add_file> <master_file> <output_file> <progress_file>")
            return
        
        add_file_path = sys.argv[1]
        master_file_path = sys.argv[2]
        output_file_path = sys.argv[3]
        progress_file = sys.argv[4]

        if not output_file_path.endswith('.xlsx'):
            output_file_path += '.xlsx'

        report_progress(progress_file, "Reading files", 10)
        # Read the input files
        add_df = read_file_with_fallback(add_file_path)
        master_df = read_file_with_fallback(master_file_path)

        report_progress(progress_file, "Transforming data", 50)
        # Transform and match the data
        transformed_df, not_found_indices = transform_data(add_df, master_df)

        report_progress(progress_file, "Saving transformed file", 80)
        # Save the output to a new Excel file
        try:
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                transformed_df.to_excel(writer, index=False)
        except Exception as e:
            print(f"Error while saving Excel file: {e}")
            traceback.print_exc()
            return

        #Adjust column widths
        adjust_column_widths(output_file_path)

        # Highlight missing rows
        highlight_missing_rows(output_file_path, not_found_indices)

        report_progress(progress_file, "Transformation completed successfully", 100)
        print(f"Transformation completed. Output saved to {output_file_path}")

    except Exception as e:
        report_progress(progress_file, f"Error: {e}", 0)
        print(f"Error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()